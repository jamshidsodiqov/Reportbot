from aiogram import types
from aiogram.dispatcher.filters.builtin import CommandStart
from aiogram.dispatcher import FSMContext
from loader import dp, bot

import openpyxl
from openpyxl import load_workbook, Workbook
import warnings
from datetime import datetime
import os
import random
from openpyxl.styles import Font, Alignment
import shutil

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Define save directories
USER_DOCUMENT_DIR = os.path.join(BASE_DIR, 'projectFiles', 'userDocument')
EDITED_USER_FILE_DIR = os.path.join(BASE_DIR, 'projectFiles', 'editedUserFile')
SUB_FILES_DIR = os.path.join(BASE_DIR, 'projectFiles', 'subFiles')
RESULT_FILE_DIR = os.path.join(BASE_DIR, 'projectFiles', 'resultFile')
EMPTY_REPORT_DIR = os.path.join(BASE_DIR, 'projectFiles', 'empty_report')

# Ensure directories exist
os.makedirs(USER_DOCUMENT_DIR, exist_ok=True)
os.makedirs(EDITED_USER_FILE_DIR, exist_ok=True)
os.makedirs(SUB_FILES_DIR, exist_ok=True)
os.makedirs(RESULT_FILE_DIR, exist_ok=True)
os.makedirs(EMPTY_REPORT_DIR, exist_ok=True)

@dp.message_handler(CommandStart())
async def bot_start(message: types.Message):
    await message.answer(f"Hello, {message.from_user.full_name}! \n"
                         f"Send me the original PBA statistics in an Excel file.")
    await fileData.file.set()  # Set the state to indicate we are expecting a file

@dp.message_handler(content_types=types.ContentType.DOCUMENT, state=fileData.file)
async def handle_document(message: types.Message, state: FSMContext):
    if message.document and message.document.file_name.endswith('.xlsx'):

# Download and save the file
        file_info = await bot.get_file(message.document.file_id)
        file_name = message.document.file_name
        file_path = file_info.file_path
        downloaded_file = await bot.download_file(file_path)

        # Define the directory and file path to save the file
        save_path = os.path.join(USER_DOCUMENT_DIR, message.document.file_name)

        # Save the file
        with open(save_path, 'wb') as new_file:
            new_file.write(downloaded_file.getvalue())

        await message.answer("Thank you for sending the Excel file. Processing...")


# Logic part for edit excel file:

        # Option to suppress specific warnings
        warnings.filterwarnings("ignore", category=UserWarning)

        # Load the workbook and select the active worksheet
        book = load_workbook(save_path)
        sheet = book.active

        # Create a new workbook and copy data to ensure styles are consistent
        new_book = Workbook()
        new_sheet = new_book.active

        # Copy the header row to the new sheet
        header = [cell.value for cell in sheet[1]]
        new_sheet.append(header)

        # Copy data and apply styles
        for row in sheet.iter_rows(values_only=True):
            new_sheet.append(row)

        # Center align all cells
        for row in new_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

        # Delete unnecessary columns in reverse order
        need_delete_columns = [1, 2, 4, 5, 6, 7, 10, 12, 14, 15, 18, 19, 21]
        need_delete_columns.sort(reverse=True)
        for col in need_delete_columns:
            new_sheet.delete_cols(col)

        # Change column width
        column_widths = {'A': 25, 'B': 25, 'C': 25, 'D': 25, 'E': 45, 'F': 25, 'G': 25, 'H': 25}
        for col, width in column_widths.items():
            new_sheet.column_dimensions[col].width = width

        # Define the target values to keep
        target_values = [
            'Fault stop',
            'Tower base stop',
            'Service mode',
            'Periodic service stop',
            'HMI stop',
            # 'Description of running status word'
        ]

        # Collect rows to delete
        rows_to_delete = []

        for row in new_sheet.iter_rows(min_row=2, max_col=new_sheet.max_column,
                                       max_row=new_sheet.max_row):  # Skip header
            cell = row[4]  # Get the cell in column 'E'
            if cell.value is None or cell.value not in target_values:
                rows_to_delete.append(cell.row)

        # Delete rows in reverse order to avoid shifting issues
        for row in reversed(rows_to_delete):
            new_sheet.delete_rows(row)

        new_sheet.delete_rows(1)
        # Save the new workbook
        edited_save_path = os.path.join(EDITED_USER_FILE_DIR, 'editedPBA.xlsx')
        new_book.save(edited_save_path)

# Send result file and edited:
        await message.answer("Sending you the edited PBA statistics document...")

        # Check if the file exists
        if os.path.exists(edited_save_path):
            try:
                # Send the file to the user
                await bot.send_document(message.chat.id, types.InputFile(edited_save_path))
            except Exception as e:
                await message.answer("There was an error sending the file. Please try again later.")
        else:
            await message.answer("The requested file does not exist.")

        await message.answer("Wait a second. 'SEPCOIII Zarafshan Wind Power Project O&M Daily Report' file is under processing...")

#Clean all project files except original file

        # Define the folder containing the files
        Folder_path = [
            SUB_FILES_DIR,
            RESULT_FILE_DIR,
            USER_DOCUMENT_DIR,
            #EDITED_USER_FILE_DIR
        ]
        # Ensure the folder exists
        for folder_path in Folder_path:
            if os.path.exists(folder_path):
                # Iterate over all the files in the folder
                for file_name in os.listdir(folder_path):
                    # Construct the full file path
                    file_path = os.path.join(folder_path, file_name)

                    # Check if it's a file (not a directory)
                    if os.path.isfile(file_path):
                        try:
                            os.remove(file_path)  # Delete the file
                        except Exception as e:
                            pass

    else:
        await message.answer("Please send a valid Excel file (.xlsx).")