import openpyxl
from aiogram import types
from aiogram.dispatcher.filters.builtin import CommandStart
from aiogram.dispatcher import FSMContext
from loader import dp, bot
from datetime import date, timedelta
from states.userStates import fileData
import os
import pandas as pd
import shutil
import random
from openpyxl import load_workbook
from collections import defaultdict

from openpyxl.styles import Side, Border, Alignment, Font, PatternFill


BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Define save directories
USER_DOCUMENT_DIR = os.path.join(BASE_DIR, 'projectFiles', 'userDocument')
EDITED_USER_FILE_DIR = os.path.join(BASE_DIR, 'projectFiles', 'editedUserFile')
RESULT_FILE_DIR = os.path.join(BASE_DIR, 'projectFiles', 'resultFile')
EMPTY_REPORT_DIR = os.path.join(BASE_DIR, 'projectFiles', 'empty_report')

# Ensure directories exist
os.makedirs(USER_DOCUMENT_DIR, exist_ok=True)
os.makedirs(EDITED_USER_FILE_DIR, exist_ok=True)
os.makedirs(RESULT_FILE_DIR, exist_ok=True)
os.makedirs(EMPTY_REPORT_DIR, exist_ok=True)

@dp.message_handler(CommandStart())
async def bot_start(message: types.Message):
    await message.answer(f"Hello, {message.from_user.full_name}! \n"
                         f"Send me the original PBA statistics in an Excel file.")
    await fileData.excelFile.set()  # Set the state to indicate we are expecting a file

@dp.message_handler(content_types=types.ContentType.DOCUMENT, state=fileData.excelFile)
async def handle_document(message: types.Message, state: FSMContext):
    if message.document and message.document.file_name.endswith('.xlsx'):

        file_info = await bot.get_file(message.document.file_id)
        file_name = message.document.file_name
        file_path = file_info.file_path
        downloaded_file = await bot.download_file(file_path)

        # Define the directory and file path to save the file
        save_path = os.path.join(USER_DOCUMENT_DIR, message.document.file_name)

        # Save the file
        with open(save_path, 'wb') as new_file:
            new_file.write(downloaded_file.getvalue())

        await message.answer("Thank you for sending the Excel file. Processing...")

        import warnings
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

        await state.finish()

        dr = pd.read_excel(save_path, header=None)  # Read without assuming any header
        dr = dr.iloc[2:].reset_index(drop=True)  # Remove first two rows
        dr.columns = dr.iloc[0]  # Set the first row as header
        dr = dr.drop(0).reset_index(drop=True)  # Drop the new header row from data
        dr.to_excel(save_path, index=False)

        # Edit file using Pandas:

        df_SAVE = pd.read_excel(save_path)

        cols = [
            "Device Name",
            "Starting time",
            "End Time",
            "Duration (m)",
            "Description of running status word",
            "error code",
            "Fault description",
            "Lost power generation (kWh)",
        ]

        new_df = df_SAVE[cols][df_SAVE['Description of running status word'].isin(['Fault stop',
                                                                                   'Tower base stop',
                                                                                   'Tower base emergency stop',
                                                                                   'Service mode',
                                                                                   'Periodic service stop',
                                                                                   'HMI stop',
                                                                                   'Nacelle stop', ])]

        # Save the file
        file_path = os.path.join(EDITED_USER_FILE_DIR, 'new_pba.xlsx')
        new_df.to_excel(file_path, index=False)

        # MAKE FINAL REPORT FILE:

        # Assuming EDITED_USER_FILE_DIR and BASE_DIR are already defined
        file_path = os.path.join(EDITED_USER_FILE_DIR, 'new_pba.xlsx')

        # Read the file
        df = pd.read_excel(file_path)

        df['Starting time'] = df['Starting time'].dt.strftime('%H:%M')
        df['End Time'] = df['End Time'].dt.strftime('%H:%M')
        data_list = df.values.tolist()

        # Using a defaultdict to group the data
        grouped_data = defaultdict(lambda: {
            'min_time': None,
            'max_time': None,
            'descriptions': set(),
            'total_power': 0,
            'total_duration': 0,
            'fault_details': []  # List to store fault number and description
        })

        # Processing the data
        for row in data_list:
            turbine, start_time, end_time, duration, description, fault_num, fault_desc, lost_power = row
            entry = grouped_data[turbine]

            # Update min and max time
            entry['min_time'] = min(start_time, entry['min_time']) if entry['min_time'] else start_time
            entry['max_time'] = max(end_time, entry['max_time']) if entry['max_time'] else end_time

            # Collect unique descriptions
            entry['descriptions'].add(description)

            # If description is 'Fault stop', append fault details
            if description == 'Fault stop' and fault_num != '--' and fault_desc != '--':
                # Append fault details without adding 'Fault stop' again
                entry['fault_details'].append([fault_num, fault_desc])

            # Sum up power losses and durations
            entry['total_power'] += lost_power
            entry['total_duration'] += duration

        # Preparing the result list
        result = []
        for turbine, data in grouped_data.items():
            descriptions = ', '.join(data['descriptions'])

            # If 'Fault stop' is present, include fault details
            if 'Fault stop' in data['descriptions']:
                fault_details = data['fault_details'][0] # Assume there's only one fault stop per turbine
                descriptions = list(data['descriptions'])

                # Split the fault details into separate list entries
                fault_num = fault_details[0]  # Fault number
                fault_desc = fault_details[1]  # Fault description
            else:
                fault_num = ''
                fault_desc = ''

            # Create the result entry with separate columns for fault number and fault description
            result.append([
                turbine,
                data['min_time'],
                data['max_time'],
                data['total_duration'],  # Include total duration
                descriptions,
                fault_num,  # Fault number as a separate entry
                fault_desc,  # Fault description as a separate entry
                f"{random.randint(2, 5)} Engineers",  # Random engineer count
                data['total_power']
            ])

        result = sorted(result, key=lambda x: x[1])
        for i in result:
            if i[8] == 0.0:
                result.remove(i)

        size = len(result)

        # COPY ORIGINAL EMPTY REPORT FILE TO RESULT FILE FOLDER

        # Assuming BASE_DIR, EMPTY_REPORT_DIR, and RESULT_FILE_DIR are already defined
        source_file_path = os.path.join(EMPTY_REPORT_DIR, 'original_report.xlsx')
        destination_file_path = os.path.join(RESULT_FILE_DIR, 'report.xlsx')

        # Copy the file
        shutil.copy(source_file_path, destination_file_path)

        # Load the existing workbook
        wb = load_workbook(destination_file_path)
        sheet = wb.active

        Analys = []
        for item in result:
            Analys.append([item[0], item[4]])

        grouped = {}
        for item in Analys:
            group = item[1]
            turbine = item[0]
            if 'Fault stop' in group:
                group = 'Fault stop'
            else:
                group = 'Service mode'
            if group not in grouped:
                grouped[group] = []
            grouped[group].append(turbine)
        x = ''
        for group, turbines in grouped.items():
            turbines_str = ', '.join(turbines)
            x += (f"{turbines_str} - {group}\n")
        sheet['H17'] = x

        sheet.insert_rows(19,19+size*2)

        for row_num in range(19, 19 + size * 2):
            sheet.row_dimensions[row_num].height = 48

        for i in range(19, 19 + size * 2 + 2):
            sheet.merge_cells(f'C{i}:L{i}')

        # Define a style for the outside border
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Apply the border   to the desired range of cells
        for row in sheet.iter_rows(min_row=19, max_row=19 + size * 2 + 1, min_col=1, max_col=12):
            for idx, cell in enumerate(row, start=1):  # Use idx to represent the column number (starting from 1)
                cell.border = thin_border

                cell.font = Font(size=16, name='等线')

                # Apply different alignment based on column index
                if idx == 1 or idx == 2:  # If it is the first or second column
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if idx == 1:
                        cell.font = Font(bold=True, size=16, name='等线')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # Starting row and column indices
        start_row = 19
        time_column = 2  # Column B (Excel index)
        main_column = 3  # Column C (Excel index)
        total_lost_power = 0
        planned_lost_power = 0
        planned_duration = 0
        Inexcusable_lost_power = 0
        Inexcusable_duration = 0
        num = 1

        sheet[f'A{start_row + size * 2}'] = 'NO.'
        sheet[f'B{start_row + size * 2}'] = 'Time'
        sheet[f'C{start_row + size * 2}'] = 'Incidents\Accidents Records'

        sheet[f'A{start_row + size * 2 + 1}'] = '1'
        sheet[f'B{start_row + size * 2 + 1}'] = '/'
        sheet[f'C{start_row + size * 2 + 1}'] = '/'

        # Define a blue fill style
        blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

        target_row = [start_row + size * 2, start_row + size * 2 + 1]

        # Loop through the specific row (target_row) to apply formatting
        for cell in ['A', 'B', 'C']:
            for target in target_row:
                cell_ref = sheet[f'{cell}{target}']
                cell_ref.font = Font(bold=True, name='Times New Roman', size=18)
                cell_ref.alignment = Alignment(horizontal='center', vertical='center')
                if target == start_row + size * 2:
                    cell_ref.fill = blue_fill

        fault_time = 0
        fault_power = 0

        for i in range(0, size):

            if result[i][8] != 0.0:
                sheet[f'A{start_row}'] = num
                sheet[f'B{start_row}'] = result[i][1]
                duration = round(result[i][3] / 60, 2)
                if 'Fault stop' in result[i][4]:
                    for d in data_list:
                        if result[i][5] in d:
                            sheet[f'C{start_row}'] = f"WTG{result[i][0][1:]}, {result[i][4][0]} to {result[i][4][1]}, {result[i][5]}, {result[i][6]}, {duration}h, {result[i][7]}"
                else:
                    sheet[f'C{start_row}'] = f"WTG{result[i][0][1:]}, {result[i][4]},{duration}h, {result[i][7]}"

                total_lost_power += result[i][8]

                if 'Fault stop' in result[i][4]:
                    for d in data_list:
                        if result[i][5] in d:
                            Inexcusable_lost_power += d[7]
                            Inexcusable_duration += round(d[3]/60,2)
                            fault_time = round(d[3]/60,2)
                            fault_power = d[7]
                if 'Service mode' in result[i][4]:
                    planned_lost_power += result[i][8]
                    planned_duration += duration

                sheet[f'B{start_row + 1}'] = result[i][2]
                sheet[f'C{start_row + 1}'] = f"WTG{result[i][0][1:]} resume power generation"
                sheet[f'A{start_row + 1}'] = num + 1
                start_row += 2
                num += 2

        today = date.today()

        sheet['G17'] = round(total_lost_power / 1000, 3)
        sheet['F17'] = f'{round(planned_duration - fault_time, 2)} / {round((planned_lost_power - fault_power )/ 1000, 3)}'
        sheet['E17'] = f'{round(Inexcusable_duration, 2)} / {round(Inexcusable_lost_power / 1000, 3)}'
        sheet['K2'] = today - timedelta(days=1)

        # Save the workbook
        edited_save_path = os.path.join(RESULT_FILE_DIR,
                                        f'SEPCOIII Zarafshan Daily report {today - timedelta(days=1)}.xlsx')
        wb.save(edited_save_path)

        # Send result file and edited:
        await message.answer(f"Sending SEPCOIII Zarafshan Daily report {today - timedelta(days=1)} document...")

        # Check if the file exists
        if os.path.exists(edited_save_path):
            try:
                # Send the file to the user
                await bot.send_document(message.chat.id, types.InputFile(edited_save_path))
            except Exception as e:
                await message.answer("There was an error sending the file. Please try again later.")
        else:
            await message.answer("The requested file does not exist.")

        # Clean all project files except original file

        Folder_path = [
            RESULT_FILE_DIR,
            USER_DOCUMENT_DIR,
            EDITED_USER_FILE_DIR
        ]
        # Ensure the folder exists
        for folder_path in Folder_path:
            if os.path.exists(folder_path):
                # Iterate over all the files in the folder
                for file_name in os.listdir(folder_path):
                    # Construct the full file path
                    file_path = os.path.join(folder_path, file_name)
                    # Check if it's a file (not a directory)
                    if os.path.isfile(file_path):
                        try:
                            os.remove(file_path)  # Delete the file
                        except Exception as e:
                            pass

    else:
        await message.answer("Please send a valid Excel file (.xlsx).")