import pandas as pd
from collections import defaultdict
import random
from openpyxl import load_workbook
from copy import copy
from collections import defaultdict

from openpyxl.styles import Side, Border, Alignment, Font, PatternFill

# Reading the data
df = pd.read_excel('new_pba.xlsx')
df['Starting time'] = df['Starting time'].dt.strftime('%H:%M')
df['End Time'] = df['End Time'].dt.strftime('%H:%M')
data_list = df.values.tolist()

# Using a defaultdict to group the data
grouped_data = defaultdict(lambda: {
    'min_time': None,
    'max_time': None,
    'descriptions': set(),
    'total_power': 0,
    'total_duration': 0,
    'fault_details': []  # List to store fault number and description
})

# Processing the data
for row in data_list:
    turbine, start_time, end_time, duration, description, fault_num, fault_desc, lost_power = row
    entry = grouped_data[turbine]

    # Update min and max time
    entry['min_time'] = min(start_time, entry['min_time']) if entry['min_time'] else start_time
    entry['max_time'] = max(end_time, entry['max_time']) if entry['max_time'] else end_time

    # Collect unique descriptions
    entry['descriptions'].add(description)

    # If description is 'Fault stop', append fault details
    if description == 'Fault stop' and fault_num != '--' and fault_desc != '--':
        # Append fault details without adding 'Fault stop' again
        entry['fault_details'].append([fault_num, fault_desc])

    # Sum up power losses and durations
    entry['total_power'] += lost_power
    entry['total_duration'] += duration

# Preparing the result list
result = []
for turbine, data in grouped_data.items():
    descriptions = ', '.join(data['descriptions'])

    # If 'Fault stop' is present, include fault details
    if 'Fault stop' in data['descriptions']:
        fault_details = data['fault_details'][0]  # Assume there's only one fault stop per turbine
        descriptions = 'Fault stop'  # Just the 'Fault stop' description

        # Split the fault details into separate list entries
        fault_num = fault_details[0]  # Fault number
        fault_desc = fault_details[1]  # Fault description
    else:
        fault_num = ''
        fault_desc = ''

    # Create the result entry with separate columns for fault number and fault description
    result.append([
        turbine,
        data['min_time'],
        data['max_time'],
        data['total_duration'],  # Include total duration
        descriptions,
        fault_num,  # Fault number as a separate entry
        fault_desc,  # Fault description as a separate entry
        f"{random.randint(2, 5)} Engineers",  # Random engineer count
        data['total_power']
    ])

result = sorted(result, key=lambda x: x[1])
size = len(result)

for row in result:
    print(row)

#Load the existing workbook
wb = load_workbook('report.xlsx')
sheet = wb.active  # Select active sheet or specify the name: wb['SheetName']

Analys = []
for item in result:
    Analys.append([item[0], item[4]])

grouped = {}
for item in Analys:
    group = item[1]
    turbine = item[0]
    if 'Service mode' in group or 'Nacelle stop' in group:
        group = 'Service mode'
    if group not in grouped:
        grouped[group] = []
    grouped[group].append(turbine)
x = ''
for group, turbines in grouped.items():
    turbines_str = ', '.join(turbines)
    x +=(f"{turbines_str} - {group}\n")
sheet['H17'] = x

# sheet.insert_rows(19,19+size*2)

for row_num in range(19, 19+size*2):
    sheet.row_dimensions[row_num].height = 48

for i in range(19, 19+size*2+2):
    sheet.merge_cells(f'C{i}:L{i}')

# Define a style for the outside border
thin_border = Border(left=Side(style='thin'),
                       right=Side(style='thin'),
                       top=Side(style='thin'),
                       bottom=Side(style='thin'))

# Apply the border   to the desired range of cells
for row in sheet.iter_rows(min_row=19, max_row=19+size*2+1, min_col=1, max_col=12):
    for idx, cell in enumerate(row, start=1):  # Use idx to represent the column number (starting from 1)
        cell.border = thin_border

        cell.font = Font(size=16, name='等线')

        # Apply different alignment based on column index
        if idx == 1 or idx == 2:  # If it is the first or second column
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if idx == 1:
                cell.font = Font(bold=True, size=16, name='等线')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

# Starting row and column indices
start_row = 19
time_column = 2  # Column B (Excel index)
main_column = 3  # Column C (Excel index)
total_lost_power = 0
planned_lost_power = 0
planned_duration = 0
Inexcusable_lost_power = 0
Inexcusable_duration = 0
num = 1

sheet[f'A{start_row+size*2}'] = 'NO.'
sheet[f'B{start_row+size*2}'] = 'Time'
sheet[f'C{start_row+size*2}'] = 'Incidents\Accidents Records'

sheet[f'A{start_row+size*2+1}'] = '1'
sheet[f'B{start_row+size*2+1}'] = '/'
sheet[f'C{start_row+size*2+1}'] = '/'

# Define a blue fill style
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

target_row = [start_row + size * 2, start_row + size * 2+1]

# Loop through the specific row (target_row) to apply formatting
for cell in ['A', 'B', 'C']:
    for target in target_row:
        cell_ref = sheet[f'{cell}{target}']
        cell_ref.font = Font(bold=True, name='Times New Roman', size=18)
        cell_ref.alignment = Alignment(horizontal='center', vertical='center')
        if target == start_row + size * 2:
            cell_ref.fill = blue_fill

for i in range(0,size):
    sheet[f'A{start_row}'] = num
    sheet[f'B{start_row}'] = result[i][1]
    duration = round(result[i][3]/60,2)
    if 'Fault stop' in result[i][4]:
        sheet[f'C{start_row}'] = f"WTG{result[i][0][1:]}, {result[i][4]}, {result[i][5]}, {result[i][6]}, {duration}h, {result[i][7]}"
    else:
        sheet[f'C{start_row}'] = f"WTG{result[i][0][1:]}, {result[i][4]},{duration}h, {result[i][7]}"

    total_lost_power += result[i][8]

    if 'Fault stop' in result[i][4]:
        Inexcusable_lost_power += result[i][8]
        Inexcusable_duration += duration
    else:
        planned_lost_power += result[i][8]
        planned_duration += duration

    sheet[f'B{start_row+1}'] = result[i][2]
    sheet[f'C{start_row+1}'] = f"WTG{result[i][0][1:]} resume power generation"
    sheet[f'A{start_row+1}'] = num + 1
    start_row+=2
    num+=2

sheet['G17'] = round(total_lost_power/1000,3)
sheet['F17'] = f'{round(planned_duration,2)} / {round(planned_lost_power/1000,3)}'
sheet['E17'] = f'{round(Inexcusable_duration,2)} / {round(Inexcusable_lost_power/1000,3)}'


# Save the workbook
wb.save('report.xlsx')
